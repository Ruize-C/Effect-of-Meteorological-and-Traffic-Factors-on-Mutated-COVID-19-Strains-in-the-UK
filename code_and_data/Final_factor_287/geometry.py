from openpyxl import load_workbook
import numpy as np
import pandas as pd
from sqlalchemy import false
area=[]
file_1=load_workbook(r"BNU-MathModeling-2022\Final_factor_287\filenames.xlsx")
sheet_1=file_1.worksheets[0]
for row in sheet_1.rows:
    if row[0].value!="LAD20NM":
        area.append(row[0].value)
file_2=load_workbook(r"BNU-MathModeling-2022\Final_factor_287\Wards__December_2016__Boundaries_UK_BGC.xlsx")
sheet_2=file_2.worksheets[0]
# file_3=load_workbook(r"BNU-MathModeling-2022\Final_factor_287\all_save_data\Adur.csv")
# sheet_3=file_3.worksheets[0]

lantitude=[]
longitude=[]
# area_0=[]
for x in area:
    lo=[]
    la=[]
    for row in sheet_2.rows:
        # if row[5].value!="LAD16NM":
        #     area_0.append(row[5].value)
        if row[5].value == x:
            lo.append(row[8].value)
            la.append(row[9].value)
    longitude.append(np.mean(lo))
    lantitude.append(np.mean(la))

# area_0=set(area_0)
# n=0
# for x in area:
#     if x in area_0:
#         n+=1

data={"area":pd.Series(area),"longitude":pd.Series(longitude),"langitude":pd.Series(lantitude)}
df=pd.DataFrame(data)
print(df)
df.to_csv(r"BNU-MathModeling-2022\Final_factor_287\location.csv",index=false)